import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule
import pandas as pd
import os
import re
from io import BytesIO



def style_excel_workbook(file_stream):
    wb = openpyxl.load_workbook(file_stream)
    ws = wb.active  # Assuming your data is in the first (active) worksheet

    max_row = ws.max_row
    max_col = ws.max_column

    # ---------------------------------------------------------
    # 1. Apply a thick border only to the outside edges
    # ---------------------------------------------------------
    thick_side = Side(border_style='thick', color='000000')
    
    # Top & bottom edges
    for col in range(1, max_col + 1):
        top_cell = ws.cell(row=1, column=col)
        bottom_cell = ws.cell(row=max_row, column=col)
        top_cell.border = Border(top=thick_side)
        bottom_cell.border = Border(bottom=thick_side)

    # Left & right edges
    for row in range(1, max_row + 1):
        left_cell = ws.cell(row=row, column=1)
        right_cell = ws.cell(row=row, column=max_col)
        left_cell.border = Border(left=thick_side)
        right_cell.border = Border(right=thick_side)

    # ---------------------------------------------------------
    # 2. Center align columns E through O
    # ---------------------------------------------------------
    # Column E=5, Column O=15 in 1-based indexing
    for col in range(5, 16):
        for row in range(1, max_row + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # ---------------------------------------------------------
    # 3. Reverse color scale on “Total Opens” (column O => col 15)
    #    so 0 is Red, 10 is Green
    # ---------------------------------------------------------
    color_scale_rule = ColorScaleRule(
        start_type='num', start_value=0, start_color='F8696B',  # Red
        mid_type='num',   mid_value=5,   mid_color='FFEB84',    # Yellow
        end_type='num',   end_value=10,  end_color='63BE7B'     # Green
    )
    data_range = f"O2:O{max_row}"  # from row 2 to the last row
    ws.conditional_formatting.add(data_range, color_scale_rule)

    # ---------------------------------------------------------
    # 4. Mailchimp vs. Constant Contact in ESP (column P => col 16)
    #    with black text, different fill colors
    # ---------------------------------------------------------
    # We'll also set column P wider (e.g. 30) so text isn't cut off
    ws.column_dimensions["P"].width = 30

    for row in range(2, max_row + 1):
        esp_cell = ws.cell(row=row, column=16)  # “ESP” column
        esp_value = esp_cell.value
        if esp_value == 'Mailchimp':
            # Yellow fill, black bold text
            esp_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            esp_cell.font = Font(color='000000', bold=True)
            esp_cell.alignment = Alignment(horizontal='center', vertical='center')
        elif esp_value == 'Constant Contact':
            # Blue fill, black bold text
            esp_cell.fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
            esp_cell.font = Font(color='000000', bold=True)
            esp_cell.alignment = Alignment(horizontal='center', vertical='center')

    # ---------------------------------------------------------
    # 5. Make rows uniformly shorter
    # ---------------------------------------------------------
    # Example: set row height to 14
    for row in range(1, max_row + 1):
        ws.row_dimensions[row].height = 14

    # ---------------------------------------------------------
    # 6. Add filters & sort by “Total Opens” descending
    # ---------------------------------------------------------
    ws.auto_filter.ref = ws.dimensions

    data = list(ws.iter_rows(values_only=True))
    headers = data[0]

    # If there's a header named "Total Opens," find its index
    # Column O is index 14 zero-based
    if "Total Opens" in headers:
        total_opens_idx = headers.index("Total Opens")
    else:
        total_opens_idx = 14

    # Sort by total opens descending
    sorted_data = sorted(
        data[1:], 
        key=lambda x: x[total_opens_idx] if x[total_opens_idx] else 0,
        reverse=True
    )

    # Clear existing data from row 2 down
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.value = None

    # Rewrite sorted data
    for row_idx, row_data in enumerate(sorted_data, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell_value)

    # ---------------------------------------------------------
    # Save
    # ---------------------------------------------------------
    file_stream.seek(0)
    wb.save(file_stream)
    file_stream.seek(0)
    return file_stream


# ----------------------------
# 1. Mailchimp Processing
# ----------------------------
def process_email_opens_for_vips(vip_list_path, email_openers_directory):
    import os
    import pandas as pd
    
    # Load the VIP list
    vip_df = pd.read_excel(vip_list_path)
    vip_df.set_index('Email Address', inplace=True)
    
    # We fix 10 campaigns (Campaign 1..10) for Mailchimp just like CC
    campaign_names = [f"Campaign {i}" for i in range(1, 11)]
    for name in campaign_names:
        vip_df[name] = ''  # Initialize empty
    
    # Process each email opener file found in the directory
    opener_files_processed = 0
    for file in sorted(os.listdir(email_openers_directory)):
        if file.endswith('-blast.csv') and opener_files_processed < len(campaign_names):
            file_path = os.path.join(email_openers_directory, file)
            openers_df = pd.read_csv(file_path)
            
            # Use the next "Campaign #" column
            current_campaign = campaign_names[opener_files_processed]
            
            # Mark opens with an 'X'
            for email in openers_df['Email Address']:
                if email in vip_df.index:
                    vip_df.loc[email, current_campaign] = 'X'
            
            opener_files_processed += 1
    
    # Calculate total opens (across these 10 campaigns)
    vip_df['Total Opens'] = vip_df[campaign_names].apply(lambda row: (row == 'X').sum(), axis=1)
    
    # Reset index so "Email Address" is visible in final
    vip_df.reset_index(inplace=True)
    
    # Add the ESP column
    vip_df['ESP'] = 'Mailchimp'
    
    return vip_df

# ----------------------------
# 2. Constant Contact Processing
# ----------------------------
def process_cc_email_opens_for_vips(vip_list_path, email_openers_directory):
    # Load the VIP list
    vip_df = pd.read_excel(vip_list_path)
    vip_df.set_index('Email address', inplace=True)

    # We fix 10 campaigns (Campaign 1..10)
    campaign_dates = [f"Campaign {i}" for i in range(1, 11)]
    for c in campaign_dates:
        vip_df[c] = ''  # Initialize as empty

    # Process each email opener file found in the directory
    opener_files_processed = 0
    for file in sorted(os.listdir(email_openers_directory)):
        if re.match(r'contact_export_.*\.csv', file) and opener_files_processed < len(campaign_dates):
            file_path = os.path.join(email_openers_directory, file)
            openers_df = pd.read_csv(file_path)
            current_campaign = campaign_dates[opener_files_processed]

            for email in openers_df['Email address']:
                if email in vip_df.index:
                    vip_df.loc[email, current_campaign] = 'X'
            opener_files_processed += 1

    # Calculate total opens for each VIP
    vip_df['Total Opens'] = vip_df[campaign_dates].apply(lambda row: (row == 'X').sum(), axis=1)
    
    # Reset the index to include "Email address" in the output
    vip_df.reset_index(inplace=True)

    #
    #  >>>>>>>>>>>>>  THE KEY CHANGE: RENAME THE CC COLUMNS  <<<<<<<<<<<<<<
    #
    vip_df.rename(
        columns={
            "Email address": "Email Address",  # match Mailchimp
            "First name": "First Name",
            "Last name": "Last Name",
            "Tags": "TAGS"  # only do this if your CC VIP list has a column "Tags" you want to become "TAGS"
        },
        inplace=True
    )

    # Add the ESP column
    vip_df['ESP'] = 'Constant Contact'

    return vip_df

# ----------------------------
# 3. Streamlit App Interface
# ----------------------------
def main():
    st.title("VIP Report Generator")
    st.markdown(
        """
        This app lets you upload:
        1. A VIP list (Excel file).
        2. The corresponding opener CSV files (up to 10).
        
        It will generate an Excel file showing who opened each campaign.
        """
    )

    # --- MAILCHIMP SECTION ---
    st.header("Mailchimp VIP Processing")
    vip_mc_file = st.file_uploader("Upload **MCVIP Excel**", type=["xlsx"], key="mcvip")
    mc_opener_files = st.file_uploader(
        "Upload Mailchimp **opener** CSVs (up to 10). Filenames must end with `-blast.csv`",
        type=["csv"],
        accept_multiple_files=True,
        key="mc_opener_files"
    )

    if st.button("Generate Mailchimp VIP Report"):
        if vip_mc_file and mc_opener_files:
            mc_dir = "temp_mc_opener_files"
            os.makedirs(mc_dir, exist_ok=True)

            # Save the uploaded VIP file
            vip_mc_path = os.path.join(mc_dir, "MCVIP.xlsx")
            with open(vip_mc_path, "wb") as f:
                f.write(vip_mc_file.getbuffer())

            # Save each opener CSV
            for fobj in mc_opener_files:
                file_path = os.path.join(mc_dir, fobj.name)
                with open(file_path, "wb") as out_f:
                    out_f.write(fobj.getbuffer())

            # Process
            report_df = process_email_opens_for_vips(vip_mc_path, mc_dir)
            st.session_state["mailchimp_report"] = report_df

            # Download
            output = BytesIO()
            report_df.to_excel(output, index=False)
            output = style_excel_workbook(output)

            st.download_button(
                label="Download Mailchimp VIP Report",
                data=output,
                file_name="Email_Opens_Report_for_VIPs.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("Please upload a VIP Excel file and at least one Mailchimp opener CSV.")


    # --- CONSTANT CONTACT SECTION ---
    st.header("Constant Contact VIP Processing")
    vip_cc_file = st.file_uploader("Upload **CCVIP Excel**", type=["xlsx"], key="ccvip")
    cc_opener_files = st.file_uploader(
        "Upload Constant Contact opener CSVs (up to 10). Filenames must match `contact_export_.*.csv`",
        type=["csv"],
        accept_multiple_files=True,
        key="cc_opener_files"
    )

    if st.button("Generate Constant Contact VIP Report"):
        if vip_cc_file and cc_opener_files:
            cc_dir = "temp_cc_opener_files"
            os.makedirs(cc_dir, exist_ok=True)

            # Save the uploaded VIP file
            vip_cc_path = os.path.join(cc_dir, "CCVIP.xlsx")
            with open(vip_cc_path, "wb") as f:
                f.write(vip_cc_file.getbuffer())

            # Save each uploaded opener CSV
            for fobj in cc_opener_files:
                file_path = os.path.join(cc_dir, fobj.name)
                with open(file_path, "wb") as out_f:
                    out_f.write(fobj.getbuffer())

            # Process
            report_df = process_cc_email_opens_for_vips(vip_cc_path, cc_dir)
            st.session_state["cc_report"] = report_df

            # Download
            output = BytesIO()
            report_df.to_excel(output, index=False)
            output = style_excel_workbook(output)
            
            st.download_button(
                label="Download CC VIP Report",
                data=output,
                file_name="CC_Email_Opens_Report_for_VIPs.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("Please upload a CC VIP Excel file and at least one Constant Contact opener CSV.")


    # --- COMBINE BOTH INTO A MASTER FILE ---
    st.header("Combine Mailchimp + Constant Contact Reports (Optional)")
    st.markdown("This will create one master file if both sets of data are available.")

    if st.button("Generate Master File"):
        mc_df = st.session_state.get("mailchimp_report")
        cc_df = st.session_state.get("cc_report")

        if mc_df is not None and cc_df is not None:
            # The key is that columns like "Email Address", "First Name", etc.
            # match EXACTLY in both DataFrames so they line up properly.
            master_df = pd.concat([mc_df, cc_df], ignore_index=True)

            # Provide a single download
            master_output = BytesIO()
            master_df.to_excel(master_output, index=False)
            master_output = style_excel_workbook(master_output)


            st.download_button(
                label="Download Master File (MC + CC)",
                data=master_output,
                file_name="Master_VIP_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("You need to generate both the Mailchimp and Constant Contact reports first.")


if __name__ == "__main__":
    main()
